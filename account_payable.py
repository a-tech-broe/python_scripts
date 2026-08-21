#!/usr/bin/env python3
"""
account_payable.py -- Deductions / AP workflow automation.

Replaces the manual daily loop:

    1. open the master spreadsheet
    2. find your vendors' invoices
    3. highlight them yellow
    4. copy them into a working sheet
    5. work the deductions
    6. go back to the master and turn the finished ones green

Commands
--------
    init      write a starter ap_config.json you can edit
    inspect   show sheets, detected header row, column mapping, vendor list
    pull      copy your vendors' open invoices into a working file and
              highlight those rows YELLOW in the master
    push      read the working file, and highlight every invoice you marked
              complete as GREEN in the master
    status    quick counts: open / in progress (yellow) / complete (green)

Typical day
-----------
    python account_payable.py pull                  # morning
    ... work the file in working/deductions_YYYY-MM-DD.xlsx,
        put "done" in the Status column ...
    python account_payable.py push                  # end of day

Every command that writes to the master takes a timestamped backup first
(use --no-backup to skip) and supports --dry-run.

Note: openpyxl rewrites the whole workbook on save. Charts, pivot tables and
images in the master are not preserved. Keep the master as plain data, or
work on a copy.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")


CONFIG_NAME = "ap_config.json"

YELLOW = "FFFFFF00"
GREEN = "FF92D050"
FILL_YELLOW = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")

# columns this script adds to the working file
COL_STATUS = "Status"
COL_NOTES = "Notes"
COL_DONE_ON = "Completed On"
COL_PUSHED_ON = "Pushed On"
COL_KEY = "AP_KEY"
COL_SHEET = "AP_MASTER_SHEET"
COL_ROW = "AP_MASTER_ROW"
TRACKING_COLS = (COL_KEY, COL_SHEET, COL_ROW)

DEFAULT_CONFIG = {
    "master_file": "~/Downloads/master.xlsx",
    "master_sheet": "",                 # blank = first/active sheet
    "work_dir": "working",
    "work_file_prefix": "deductions",
    "vendors": [],                      # <-- your vendors go here
    "vendor_match": "contains",         # "contains" or "exact"
    "complete_values": ["complete", "completed", "done", "yes", "y", "x", "closed", "✓"],
    "stamp_completed_date": True,       # write today's date into a master
                                        # "Completed"/"Status" column if one exists
    "columns": {
        "vendor": ["vendor", "vendor name", "vendor id", "supplier", "supplier name",
                   "payee", "customer", "customer name", "account name"],
        "invoice": ["invoice", "invoice #", "invoice no", "invoice no.", "invoice number",
                    "inv #", "inv no", "document", "document no", "doc no", "doc number",
                    "reference", "ref", "deduction #", "deduction number", "claim", "claim #",
                    "chargeback #"],
        "amount": ["amount", "invoice amount", "deduction amount", "open amount",
                   "net amount", "balance", "value", "amt"],
        "date": ["date", "invoice date", "doc date", "document date", "deduction date",
                 "posting date", "due date"],
        "completed": ["completed", "completed on", "date completed", "completion date",
                      "cleared", "resolved", "status"],
    },
}


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #
def norm(value) -> str:
    """Lower-cased, whitespace-collapsed text form of a cell value."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return re.sub(r"\s+", " ", text).lower()


def norm_key_part(value) -> str:
    """Like norm(), but also drops punctuation so INV-001 == inv 001."""
    text = norm(value)
    return re.sub(r"[^a-z0-9]", "", text)


def money(value) -> str:
    try:
        return f"{float(str(value).replace(',', '').replace('$', '')):.2f}"
    except (TypeError, ValueError):
        return norm_key_part(value)


def classify_fill(cell) -> str | None:
    """Return 'yellow', 'green', 'other' or None for a cell's background."""
    fill = cell.fill
    if fill is None or fill.patternType in (None, "none"):
        return None
    color = fill.fgColor
    if color is None or color.type != "rgb" or not color.rgb:
        return "other"
    rgb = str(color.rgb)
    if len(rgb) == 8:
        alpha, rgb = rgb[:2], rgb[2:]
        if alpha == "00":
            return None
    if len(rgb) != 6:
        return "other"
    try:
        r, g, b = (int(rgb[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return "other"
    if r > 200 and g > 180 and b < 150:
        return "yellow"
    if g > 110 and g - r > 25 and g - b > 25:
        return "green"
    if r > 245 and g > 245 and b > 245:
        return None
    return "other"


def row_color(ws, row: int, col_span: range) -> str | None:
    """Colour of a master row: the first coloured cell wins, green beats yellow."""
    seen = set()
    for col in col_span:
        tag = classify_fill(ws.cell(row=row, column=col))
        if tag:
            seen.add(tag)
    if "green" in seen:
        return "green"
    if "yellow" in seen:
        return "yellow"
    return next(iter(seen), None)


def paint_row(ws, row: int, col_span: range, fill: PatternFill) -> None:
    for col in col_span:
        ws.cell(row=row, column=col).fill = fill


def row_is_blank(ws, row: int, col_span: range) -> bool:
    return all(ws.cell(row=row, column=c).value in (None, "") for c in col_span)


# --------------------------------------------------------------------------- #
# config
# --------------------------------------------------------------------------- #
def config_path(args) -> Path:
    return Path(args.config).expanduser()


def load_config(args) -> dict:
    path = config_path(args)
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))  # deep copy
    if path.exists():
        user = json.loads(path.read_text())
        for key, value in user.items():
            if key == "columns" and isinstance(value, dict):
                cfg["columns"].update(value)
            else:
                cfg[key] = value
    elif not getattr(args, "master", None):
        sys.exit(f"No {path} found and no --master given.\n"
                 f"Run:  python {Path(sys.argv[0]).name} init")
    if getattr(args, "master", None):
        cfg["master_file"] = args.master
    if getattr(args, "sheet", None):
        cfg["master_sheet"] = args.sheet
    return cfg


# --------------------------------------------------------------------------- #
# master sheet parsing
# --------------------------------------------------------------------------- #
class MasterSheet:
    """A master worksheet plus its header row and column mapping."""

    def __init__(self, path: Path, sheet_name: str, cfg: dict):
        self.path = path
        keep_vba = path.suffix.lower() == ".xlsm"
        self.wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
        self.vals = load_workbook(path, data_only=True, read_only=False)
        if sheet_name:
            if sheet_name not in self.wb.sheetnames:
                sys.exit(f"Sheet {sheet_name!r} not in {path.name}. "
                         f"Found: {', '.join(self.wb.sheetnames)}")
            self.ws = self.wb[sheet_name]
            self.vws = self.vals[sheet_name]
        else:
            self.ws = self.wb.active
            self.vws = self.vals[self.ws.title]
        self.sheet_name = self.ws.title
        self.header_row = self._detect_header_row(cfg)
        self.headers = self._read_headers()
        self.col_span = range(1, max(self.headers) + 1) if self.headers else range(1, 2)
        self.cols = self._map_columns(cfg)

    # -- detection ---------------------------------------------------------- #
    def _detect_header_row(self, cfg: dict) -> int:
        known = {alias for aliases in cfg["columns"].values() for alias in aliases}
        best_row, best_score = 1, -1
        for row in range(1, min(self.ws.max_row, 20) + 1):
            filled, hits = 0, 0
            for col in range(1, min(self.ws.max_column, 60) + 1):
                text = norm(self.vws.cell(row=row, column=col).value)
                if not text:
                    continue
                filled += 1
                if text in known or any(text.startswith(a) or a in text for a in known):
                    hits += 1
            score = hits * 10 + filled
            if hits and score > best_score:
                best_row, best_score = row, score
        return best_row

    def _read_headers(self) -> dict[int, str]:
        headers: dict[int, str] = {}
        for col in range(1, self.ws.max_column + 1):
            text = self.vws.cell(row=self.header_row, column=col).value
            if text is not None and str(text).strip():
                headers[col] = str(text).strip()
        return headers

    def _map_columns(self, cfg: dict) -> dict[str, int]:
        mapping: dict[str, int] = {}
        used: set[int] = set()
        lowered = {col: norm(name) for col, name in self.headers.items()}
        for logical, aliases in cfg["columns"].items():
            wanted = [norm(a) for a in aliases]
            hit = next((c for c, n in lowered.items()
                        if c not in used and n in wanted), None)
            if hit is None:  # loose match: header contains an alias
                hit = next((c for c, n in lowered.items()
                            if c not in used and any(a in n for a in wanted)), None)
            if hit is not None:
                mapping[logical] = hit
                used.add(hit)
        return mapping

    # -- data --------------------------------------------------------------- #
    def value(self, row: int, col: int):
        return self.vws.cell(row=row, column=col).value

    def data_rows(self):
        for row in range(self.header_row + 1, self.ws.max_row + 1):
            if row_is_blank(self.vws, row, self.col_span):
                continue
            yield row

    def key_for(self, row: int) -> str:
        vendor = norm_key_part(self.value(row, self.cols["vendor"]))
        invoice = ""
        if "invoice" in self.cols:
            invoice = norm_key_part(self.value(row, self.cols["invoice"]))
        amount = money(self.value(row, self.cols["amount"])) if "amount" in self.cols else ""
        if invoice:
            return f"{vendor}|{invoice}|{amount}"
        return f"{vendor}|ROW{row}|{amount}"   # no invoice number: fall back to row

    def save(self, backup: bool = True) -> Path | None:
        backup_path = None
        if backup:
            backup_dir = self.path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"{self.path.stem}_{stamp}{self.path.suffix}"
            shutil.copy2(self.path, backup_path)
        self.wb.save(self.path)
        return backup_path


def vendor_matches(value, cfg: dict) -> bool:
    vendors = [norm(v) for v in cfg.get("vendors", []) if str(v).strip()]
    if not vendors:
        return True                      # no filter configured: take everything
    text = norm(value)
    if not text:
        return False
    if cfg.get("vendor_match", "contains") == "exact":
        return text in vendors
    return any(v in text or text in v for v in vendors)


def require_columns(master: MasterSheet) -> None:
    if "vendor" not in master.cols:
        sys.exit(
            "Could not find a vendor column.\n"
            f"Headers on row {master.header_row}: "
            f"{', '.join(master.headers.values()) or '(none)'}\n"
            f"Add the real header name to columns.vendor in {CONFIG_NAME}."
        )


# --------------------------------------------------------------------------- #
# working file
# --------------------------------------------------------------------------- #
def work_path_for(cfg: dict, explicit: str | None) -> Path:
    if explicit:
        return Path(explicit).expanduser()
    work_dir = Path(cfg.get("work_dir") or ".").expanduser()
    name = f"{cfg.get('work_file_prefix', 'deductions')}_{date.today():%Y-%m-%d}.xlsx"
    return work_dir / name


RESERVED = {COL_STATUS, COL_NOTES, COL_DONE_ON, COL_PUSHED_ON, *TRACKING_COLS}


def work_headers(master: MasterSheet) -> list[str]:
    seen: dict[str, int] = {}
    master_names = []
    for col in sorted(master.headers):
        name = master.headers[col]
        if name in RESERVED:            # don't shadow the columns this script adds
            name = f"{name} (master)"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        master_names.append(name)
    return [COL_STATUS, COL_NOTES, *master_names,
            COL_DONE_ON, COL_PUSHED_ON, *TRACKING_COLS]


def create_work_book(headers: list[str], master: MasterSheet):
    wb = Workbook()
    ws = wb.active
    ws.title = "Deductions"
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="FF1F4E78", end_color="FF1F4E78")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        width = 12 if name in TRACKING_COLS else max(11, min(28, len(name) + 6))
        ws.column_dimensions[get_column_letter(col)].width = width
    for name in TRACKING_COLS:
        ws.column_dimensions[get_column_letter(headers.index(name) + 1)].hidden = True
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    dv = DataValidation(type="list", formula1='"open,in progress,done"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A2:A5000")
    return wb, ws


def open_work_book(path: Path, headers: list[str], master: MasterSheet, fresh: bool):
    """Return (wb, ws, colmap, existing_keys). Appends to today's file if present."""
    if path.exists() and not fresh:
        wb = load_workbook(path)
        ws = wb.active
        found = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
        missing = [h for h in (*TRACKING_COLS, COL_STATUS) if h not in found]
        if missing:
            sys.exit(f"{path} does not look like a working file "
                     f"(missing {', '.join(missing)}). Use --new or --out.")
        colmap = {name: i + 1 for i, name in enumerate(found) if name}
        key_col = colmap[COL_KEY]
        keys = {str(ws.cell(row=r, column=key_col).value)
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=key_col).value}
        return wb, ws, colmap, keys
    path.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = create_work_book(headers, master)
    colmap = {name: i + 1 for i, name in enumerate(headers)}
    return wb, ws, colmap, set()


def is_complete(value, cfg: dict) -> bool:
    text = norm(value)
    if not text:
        return False
    return text in [norm(v) for v in cfg["complete_values"]]


# --------------------------------------------------------------------------- #
# commands
# --------------------------------------------------------------------------- #
def cmd_init(args) -> int:
    path = config_path(args)
    if path.exists() and not args.force:
        print(f"{path} already exists (use --force to overwrite).")
        return 0
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))
    if args.master:
        cfg["master_file"] = args.master
    cfg["vendors"] = ["ACME CORP", "GLOBEX", "INITECH"]
    path.write_text(json.dumps(cfg, indent=2) + "\n")
    print(f"Wrote {path}\n\nNext:\n"
          f"  1. set \"master_file\" to your master spreadsheet\n"
          f"  2. replace \"vendors\" with the vendors you own\n"
          f"  3. run:  python {Path(sys.argv[0]).name} inspect")
    return 0


def cmd_inspect(args) -> int:
    cfg = load_config(args)
    master = MasterSheet(Path(cfg["master_file"]).expanduser(), cfg["master_sheet"], cfg)
    print(f"File          : {master.path}")
    print(f"Sheets        : {', '.join(master.wb.sheetnames)}")
    print(f"Using sheet   : {master.sheet_name}")
    print(f"Header row    : {master.header_row}")
    print(f"Data rows     : {sum(1 for _ in master.data_rows())}")
    print("\nColumn mapping")
    for logical in cfg["columns"]:
        col = master.cols.get(logical)
        where = f"{get_column_letter(col)}  {master.headers[col]!r}" if col else "-- not found --"
        print(f"  {logical:<10} {where}")
    unmapped = [master.headers[c] for c in sorted(master.headers)
                if c not in master.cols.values()]
    if unmapped:
        print("\nOther columns : " + ", ".join(unmapped))

    counts: dict[str, int] = {}
    vendors: dict[str, list[int]] = {}
    vcol = master.cols.get("vendor")
    for row in master.data_rows():
        tag = row_color(master.ws, row, master.col_span) or "none"
        counts[tag] = counts.get(tag, 0) + 1
        if vcol:
            name = str(master.value(row, vcol) or "").strip()
            if name:
                vendors.setdefault(name, []).append(row)
    print("\nRow colours   : " + ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))

    if vendors:
        print(f"\nVendors found ({len(vendors)}) -- copy the ones you own into "
              f"\"vendors\" in {CONFIG_NAME}:")
        for name, rows in sorted(vendors.items(), key=lambda kv: -len(kv[1]))[:args.top]:
            mark = "*" if vendor_matches(name, cfg) else " "
            print(f"  {mark} {name:<45} {len(rows)} row(s)")
        if len(vendors) > args.top:
            print(f"  ... {len(vendors) - args.top} more (use --top N)")
        if cfg.get("vendors"):
            print("  (* = matches your configured vendor list)")
    return 0


def cmd_pull(args) -> int:
    cfg = load_config(args)
    master = MasterSheet(Path(cfg["master_file"]).expanduser(), cfg["master_sheet"], cfg)
    require_columns(master)
    if not cfg.get("vendors"):
        print("! No vendors configured -- pulling every open row. "
              f"Add your vendors to {CONFIG_NAME} to narrow this down.\n")

    headers = work_headers(master)
    out_path = work_path_for(cfg, args.out)
    wb, ws, colmap, existing = open_work_book(out_path, headers, master, args.new)
    if not args.new and out_path.exists():
        print(f"Appending to existing {out_path}")

    master_cols = sorted(master.headers)
    picked, skipped_done, skipped_open, dupes = [], 0, 0, 0

    for row in master.data_rows():
        if not vendor_matches(master.value(row, master.cols["vendor"]), cfg):
            continue
        color = row_color(master.ws, row, master.col_span)
        if color == "green":
            skipped_done += 1
            continue
        if color == "yellow" and not args.include_pulled:
            skipped_open += 1
            continue
        key = master.key_for(row)
        if key in existing:
            dupes += 1
            continue
        existing.add(key)
        picked.append((row, key))

    if not picked:
        print(f"Nothing new to pull. "
              f"(already complete: {skipped_done}, already pulled: {skipped_open}, "
              f"already in working file: {dupes})")
        return 0

    if args.dry_run:
        print(f"DRY RUN -- would copy {len(picked)} row(s) to {out_path} "
              f"and highlight them yellow in {master.path.name}:\n")
    else:
        for row, key in picked:
            target = ws.max_row + 1
            ws.cell(row=target, column=colmap[COL_STATUS], value="open")
            for offset, col in enumerate(master_cols):
                src = master.ws.cell(row=row, column=col)
                cell = ws.cell(row=target, column=colmap[COL_STATUS] + 2 + offset,
                               value=master.value(row, col))
                if src.number_format and src.number_format != "General":
                    cell.number_format = src.number_format
            ws.cell(row=target, column=colmap[COL_KEY], value=key)
            ws.cell(row=target, column=colmap[COL_SHEET], value=master.sheet_name)
            ws.cell(row=target, column=colmap[COL_ROW], value=row)
            paint_row(master.ws, row, master.col_span, FILL_YELLOW)

    preview(master, [r for r, _ in picked], args.limit)

    if args.dry_run:
        return 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    backup = master.save(backup=not args.no_backup)
    print(f"\nCopied {len(picked)} row(s) -> {out_path}")
    print(f"Highlighted {len(picked)} row(s) YELLOW in {master.path.name}"
          f" (sheet {master.sheet_name})")
    if skipped_done or skipped_open or dupes:
        print(f"Skipped: {skipped_done} already complete, {skipped_open} already pulled, "
              f"{dupes} already in the working file")
    if backup:
        print(f"Master backup: {backup}")
    print(f"\nWork the file, put a value from {cfg['complete_values'][:3]} in the "
          f"'{COL_STATUS}' column, then run:  "
          f"python {Path(sys.argv[0]).name} push --work \"{out_path}\"")
    return 0


def show(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value if value is not None else "")


def preview(master: MasterSheet, rows: list[int], limit: int) -> None:
    fields = [f for f in ("vendor", "invoice", "amount", "date") if f in master.cols]
    print(f"{'row':>6}  " + "  ".join(f"{f:<22}" for f in fields))
    for row in rows[:limit]:
        cells = [show(master.value(row, master.cols[f]))[:22] for f in fields]
        print(f"{row:>6}  " + "  ".join(f"{c:<22}" for c in cells))
    if len(rows) > limit:
        print(f"{'':>6}  ... and {len(rows) - limit} more")


def cmd_push(args) -> int:
    cfg = load_config(args)
    master = MasterSheet(Path(cfg["master_file"]).expanduser(), cfg["master_sheet"], cfg)
    require_columns(master)

    work_file = work_path_for(cfg, args.work)
    if not work_file.exists():
        sys.exit(f"Working file not found: {work_file}\n"
                 f"Pass it with --work, e.g. --work working/deductions_2026-01-15.xlsx")
    wwb = load_workbook(work_file, data_only=True)
    wws = wwb.active
    found = [str(wws.cell(row=1, column=c).value or "") for c in range(1, wws.max_column + 1)]
    wcol = {name: i + 1 for i, name in enumerate(found) if name}
    for needed in (COL_STATUS, COL_KEY, COL_ROW):
        if needed not in wcol:
            sys.exit(f"{work_file} has no '{needed}' column -- was it made by this script?")

    # index the master by key so rows still match after a re-sort
    index: dict[str, list[int]] = {}
    for row in master.data_rows():
        index.setdefault(master.key_for(row), []).append(row)

    done, already, unmatched, ambiguous = [], 0, [], []
    for wrow in range(2, wws.max_row + 1):
        key = wws.cell(row=wrow, column=wcol[COL_KEY]).value
        if not key:
            continue
        key = str(key)
        status = wws.cell(row=wrow, column=wcol[COL_STATUS]).value
        if not (args.all or is_complete(status, cfg)):
            continue
        if not args.force and COL_PUSHED_ON in wcol \
                and wws.cell(row=wrow, column=wcol[COL_PUSHED_ON]).value:
            already += 1
            continue

        hint = wws.cell(row=wrow, column=wcol[COL_ROW]).value
        target = None
        if isinstance(hint, int) and hint <= master.ws.max_row \
                and master.key_for(hint) == key:
            target = hint                       # row hasn't moved: fast path
        else:
            matches = index.get(key, [])
            if len(matches) == 1:
                target = matches[0]
            elif len(matches) > 1:
                ambiguous.append((wrow, key, matches))
                continue
        if target is None:
            unmatched.append((wrow, key))
            continue
        done.append((wrow, target))

    if not done:
        print(f"Nothing to push from {work_file.name}. "
              f"(already pushed: {already}, unmatched: {len(unmatched)})")
        report_problems(unmatched, ambiguous)
        return 0

    if args.dry_run:
        print(f"DRY RUN -- would mark {len(done)} row(s) GREEN in {master.path.name}:\n")
        preview(master, [m for _, m in done], args.limit)
        report_problems(unmatched, ambiguous)
        return 0

    today = date.today()
    stamp_col = master.cols.get("completed") if cfg.get("stamp_completed_date") else None
    for wrow, mrow in done:
        paint_row(master.ws, mrow, master.col_span, FILL_GREEN)
        if stamp_col:
            cell = master.ws.cell(row=mrow, column=stamp_col)
            header = norm(master.headers.get(stamp_col, ""))
            if header == "status":
                cell.value = "Complete"
            else:
                cell.value = today
                cell.number_format = "yyyy-mm-dd"

    # stamp the working file so a re-run doesn't re-report the same rows
    wsrc = load_workbook(work_file)
    wout = wsrc.active
    if COL_PUSHED_ON in wcol:
        for wrow, _ in done:
            wout.cell(row=wrow, column=wcol[COL_PUSHED_ON], value=today).number_format = "yyyy-mm-dd"
        if COL_DONE_ON in wcol:
            for wrow, _ in done:
                cell = wout.cell(row=wrow, column=wcol[COL_DONE_ON])
                if cell.value is None:
                    cell.value = today
                    cell.number_format = "yyyy-mm-dd"
        wsrc.save(work_file)

    preview(master, [m for _, m in done], args.limit)
    backup = master.save(backup=not args.no_backup)
    print(f"\nMarked {len(done)} row(s) GREEN in {master.path.name} "
          f"(sheet {master.sheet_name})")
    if already:
        print(f"Skipped {already} row(s) pushed on an earlier run (--force to redo)")
    if backup:
        print(f"Master backup: {backup}")
    report_problems(unmatched, ambiguous)
    return 0


def report_problems(unmatched, ambiguous) -> None:
    if unmatched:
        print(f"\n! {len(unmatched)} completed row(s) could not be found in the master "
              f"(invoice edited or deleted?):")
        for wrow, key in unmatched[:10]:
            print(f"    working row {wrow}: {key}")
    if ambiguous:
        print(f"\n! {len(ambiguous)} row(s) match more than one master row -- "
              f"highlight these by hand:")
        for wrow, key, rows in ambiguous[:10]:
            print(f"    working row {wrow}: {key} -> master rows {rows}")


def cmd_status(args) -> int:
    cfg = load_config(args)
    master = MasterSheet(Path(cfg["master_file"]).expanduser(), cfg["master_sheet"], cfg)
    require_columns(master)
    mine = {"open": 0, "yellow": 0, "green": 0, "other": 0}
    total = 0
    for row in master.data_rows():
        total += 1
        if not vendor_matches(master.value(row, master.cols["vendor"]), cfg):
            continue
        color = row_color(master.ws, row, master.col_span)
        mine[color if color in mine else "open" if color is None else "other"] += 1
    scope = "my vendors" if cfg.get("vendors") else "all vendors"
    print(f"{master.path.name} / {master.sheet_name} -- {total} data rows\n")
    print(f"For {scope}:")
    print(f"  not started (no fill) : {mine['open']}")
    print(f"  in progress (yellow)  : {mine['yellow']}")
    print(f"  complete    (green)   : {mine['green']}")
    if mine["other"]:
        print(f"  other fill colour     : {mine['other']}")
    return 0


# --------------------------------------------------------------------------- #
# cli
# --------------------------------------------------------------------------- #
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="account_payable.py",
        description="Automate the AP deductions pull / work / complete loop.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Daily:  pull  ->  work the file  ->  push",
    )
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--config", default=CONFIG_NAME,
                        help=f"config file (default {CONFIG_NAME})")
    common.add_argument("--master", help="master spreadsheet (overrides the config)")
    common.add_argument("--sheet", help="sheet name in the master (overrides the config)")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("init", parents=[common], help="write a starter config file")
    p.add_argument("--force", action="store_true", help="overwrite an existing config")
    p.set_defaults(func=cmd_init)

    p = sub.add_parser("inspect", parents=[common],
                       help="show sheets, columns and vendors in the master")
    p.add_argument("--top", type=int, default=40, help="how many vendors to list")
    p.set_defaults(func=cmd_inspect)

    p = sub.add_parser("pull", parents=[common],
                       help="copy my vendors' open invoices out and highlight them yellow")
    p.add_argument("--out", help="working file to write (default working/<prefix>_<today>.xlsx)")
    p.add_argument("--new", action="store_true", help="start a fresh working file, don't append")
    p.add_argument("--include-pulled", action="store_true",
                   help="also re-pull rows already highlighted yellow")
    p.add_argument("--dry-run", action="store_true", help="show what would happen, change nothing")
    p.add_argument("--no-backup", action="store_true", help="don't back the master up first")
    p.add_argument("--limit", type=int, default=25, help="rows to preview")
    p.set_defaults(func=cmd_pull)

    p = sub.add_parser("push", parents=[common],
                       help="mark everything completed in the working file green")
    p.add_argument("--work", help="working file (default working/<prefix>_<today>.xlsx)")
    p.add_argument("--all", action="store_true",
                   help="treat every row in the working file as complete")
    p.add_argument("--force", action="store_true", help="re-push rows already pushed")
    p.add_argument("--dry-run", action="store_true", help="show what would happen, change nothing")
    p.add_argument("--no-backup", action="store_true", help="don't back the master up first")
    p.add_argument("--limit", type=int, default=25, help="rows to preview")
    p.set_defaults(func=cmd_push)

    p = sub.add_parser("status", parents=[common],
                       help="counts of open / in progress / complete")
    p.set_defaults(func=cmd_status)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        return args.func(args)
    except PermissionError as exc:
        sys.exit(f"Permission denied: {exc.filename}\n"
                 f"Close the workbook in Excel and run it again.")
    except FileNotFoundError as exc:
        sys.exit(f"File not found: {exc.filename}")


if __name__ == "__main__":
    sys.exit(main())
